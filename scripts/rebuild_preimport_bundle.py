from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from copy import deepcopy
from pathlib import Path

from openpyxl import load_workbook


SITE_ROOT = Path("/Users/chengzhihao/Documents/Codex/2026-06-20-jrcedu-master")
WORK_ROOT = Path("/Users/chengzhihao/Documents/Codex/2026-06-10/new-chat/work")
BASE_BUNDLE_PATH = WORK_ROOT / "paike-finance-preimport-private-2026-06-22.json"
SCHEDULE_JSON_PATH = WORK_ROOT / "paike_calendar_sessions.json"
MAY_SALARY_PATH = Path(
    "/Users/chengzhihao/Library/Containers/com.tencent.xinWeChat/Data/Documents/"
    "xwechat_files/cicilang001_cd31/temp/drag/小班课2026.5工资表(3).xlsx"
)
PORTAL_PREIMPORT_PATH = SITE_ROOT / "portal" / "preimport-data.js"
PUBLIC_PREIMPORT_PATH = SITE_ROOT / "public" / "portal" / "preimport-data.js"
OUT_BUNDLE_PATH = WORK_ROOT / "paike-finance-preimport-private-2026-06-23.json"

TEACHER_NAME_MAP = {
    "李舒老师": "李舒",
    "何老师": "何建军",
    "曹老师": "曹德顺",
    "潘老师": "潘云贵",
    "吴（女）": "吴水琴",
}
NON_TEACHING_FINANCE_NAMES = {"周珊"}
DEPARTED_FINANCE_ONLY_NAMES = {"万盈盈"}
IGNORE_SHEETS = {"小课产值", "日常消费表", "周珊"}
DATE_RE = re.compile(r"第\s*(\d+)\s*次\s*(\d{1,2})[.月](\d{1,2})")


def text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def clean_header(value) -> str:
    return text(value).replace(" ", "").replace("\n", "")


def number(value, fallback=None):
    if value in (None, ""):
        return fallback
    try:
        return float(value)
    except Exception:
        return fallback


def to_number(value, fallback=0):
    parsed = number(value, None)
    return fallback if parsed is None else parsed


def parse_lesson_marker(value: str) -> dict:
    raw = text(value)
    if not raw:
        return {}
    match = DATE_RE.search(raw)
    if not match:
        return {}
    lesson_no = int(match.group(1))
    month = int(match.group(2))
    day = int(match.group(3))
    if not month or not day:
        return {"raw": raw, "needsReview": True, "reason": "未识别日期"}
    return {
        "raw": raw,
        "lessonNo": lesson_no,
        "date": f"2026-{month:02d}-{day:02d}",
        "needsReview": month != 5,
        "reason": "" if month == 5 else "不是5月日期",
    }


def parse_month_day(value) -> str:
    if value in (None, ""):
        return ""
    numeric = number(value, None)
    if numeric is not None:
        month = int(numeric)
        day = round((numeric - month) * 100)
        if 1 <= month <= 12 and 1 <= day <= 31:
            return f"2026-{month:02d}-{day:02d}"
    raw = text(value)
    nums = re.findall(r"\d+", raw)
    if len(nums) >= 2:
        month, day = int(nums[0]), int(nums[1])
        if 1 <= month <= 12 and 1 <= day <= 31:
            return f"2026-{month:02d}-{day:02d}"
    return ""


def normalize_teacher_name(name: str) -> str:
    return TEACHER_NAME_MAP.get(text(name), text(name))


def find_header_row(ws) -> int:
    for row_idx in range(1, min(ws.max_row, 12) + 1):
        values = [clean_header(ws.cell(row_idx, col).value) for col in range(1, min(ws.max_column, 40) + 1)]
        if "姓名" in values and "课时费" in values and "提成" in values and "计数" in values:
            return row_idx
    return 0


def find_sections(headers: list[str]) -> list[dict]:
    sections = []
    index = 0
    while index < len(headers):
        try:
            name_col = headers.index("姓名", index) + 1
        except ValueError:
            break
        fee_col = commission_col = count_col = None
        for col_idx in range(name_col + 1, len(headers) + 1):
            header = headers[col_idx - 1]
            if fee_col is None and header == "课时费":
                fee_col = col_idx
            elif fee_col is not None and commission_col is None and header == "提成":
                commission_col = col_idx
            elif commission_col is not None and header == "计数":
                count_col = col_idx
                break
        if fee_col and commission_col and count_col and count_col > commission_col + 1:
            sections.append(
                {
                    "nameCol": name_col,
                    "feeCol": fee_col,
                    "commissionCol": commission_col,
                    "dateCols": list(range(commission_col + 1, count_col)),
                }
            )
            index = count_col
        else:
            index = name_col
    return sections[:2]


def extract_summary_rows(ws) -> list[dict]:
    header_row = 2
    headers = [clean_header(ws.cell(header_row, col).value) for col in range(1, ws.max_column + 1)]
    rows = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        first_cell = text(ws.cell(row_idx, 1).value)
        teacher_name = text(ws.cell(row_idx, 2).value)
        serial_number = number(first_cell, None)
        if not teacher_name:
            continue
        if first_cell in {"合计", "总支出", "小课支出"} or teacher_name == "合计":
            break
        if serial_number is None:
            continue
        if number(teacher_name, None) is not None:
            continue
        row = {
            "row": row_idx,
            "period": "2026-05",
            "teacherName": teacher_name,
            "financeRoleType": "nonTeaching" if teacher_name in NON_TEACHING_FINANCE_NAMES else "teaching",
            "employmentStatus": "departed" if teacher_name in DEPARTED_FINANCE_ONLY_NAMES else "active",
        }
        for col_idx, header in enumerate(headers, start=1):
            if header:
                row[header] = ws.cell(row_idx, col_idx).value
        rows.append(row)
    return rows


def extract_expense_rows(ws) -> list[dict]:
    starts = []
    for col in range(1, ws.max_column + 1):
        header = text(ws.cell(6, col).value)
        if header and not header.startswith("截止到"):
            starts.append((col, header))
    rows = []
    for start_col, category in starts:
        for row_idx in range(7, ws.max_row + 1):
            name = text(ws.cell(row_idx, start_col).value)
            raw_date = ws.cell(row_idx, start_col + 1).value if start_col + 1 <= ws.max_column else None
            quantity = number(ws.cell(row_idx, start_col + 2).value if start_col + 2 <= ws.max_column else None, None)
            amount = number(ws.cell(row_idx, start_col + 3).value if start_col + 3 <= ws.max_column else None, None)
            if not name or amount is None:
                continue
            if name in {"合计", "总计"}:
                continue
            parsed_date = parse_month_day(raw_date)
            rows.append(
                {
                    "date": parsed_date,
                    "period": "2026-05",
                    "category": category,
                    "name": name,
                    "quantity": quantity,
                    "amount": round(amount, 2),
                    "sourceCell": ws.cell(row_idx, start_col).coordinate,
                    "status": "待审核",
                    "needsReview": not bool(parsed_date),
                    "reason": "" if parsed_date else "缺少有效日期",
                    "sourceSheet": ws.title,
                    "sourceWorkbook": MAY_SALARY_PATH.name,
                }
            )
    return rows


def extract_teacher_sheet(ws) -> tuple[list[dict], list[dict], list[dict], list[dict]]:
    header_row = find_header_row(ws)
    if not header_row:
        return [], [], [], []
    teacher_name = normalize_teacher_name(ws.title)
    headers = [clean_header(ws.cell(header_row, col).value) for col in range(1, ws.max_column + 1)]
    sections = find_sections(headers)
    if not sections:
        return [], [], [], []
    payout_section = sections[0]
    revenue_section = sections[1] if len(sections) > 1 else None
    payout_records = []
    revenue_records = []
    payout_attendance = []
    revenue_attendance = []

    for row_idx in range(header_row + 1, ws.max_row + 1):
        payout_name = text(ws.cell(row_idx, payout_section["nameCol"]).value)
        if payout_name in {"合计", "总合计"}:
            break
        if payout_name:
            fee = number(ws.cell(row_idx, payout_section["feeCol"]).value, None)
            commission = number(ws.cell(row_idx, payout_section["commissionCol"]).value, None)
            markers = []
            for col_idx in payout_section["dateCols"]:
                parsed = parse_lesson_marker(ws.cell(row_idx, col_idx).value)
                if parsed:
                    parsed["cell"] = ws.cell(row_idx, col_idx).coordinate
                    markers.append(parsed)
            payout_records.append(
                {
                    "period": "2026-05",
                    "teacherName": teacher_name,
                    "teacherSheet": ws.title,
                    "studentName": payout_name,
                    "courseFee": fee,
                    "commission": commission,
                    "amount": commission,
                    "sourceSection": "payout",
                    "sourceRow": row_idx,
                    "lessonMarkerCount": len(markers),
                    "rawMarkers": [marker["raw"] for marker in markers],
                }
            )
            for marker in markers:
                payout_attendance.append(
                    {
                        "teacherName": teacher_name,
                        "teacher": teacher_name,
                        "studentName": payout_name,
                        "courseFee": fee,
                        "commission": commission,
                        "amount": commission,
                        "date": marker.get("date", ""),
                        "period": marker.get("date", "")[:7] if marker.get("date") else "",
                        "lessonNo": marker.get("lessonNo"),
                        "raw": marker.get("raw", ""),
                        "cell": marker.get("cell", ""),
                        "status": "待对账" if not marker.get("needsReview") else "待复核",
                        "needsReview": marker.get("needsReview", False),
                        "reason": marker.get("reason", ""),
                        "sourceSection": "payout",
                        "sourceWorkbook": MAY_SALARY_PATH.name,
                        "sourceSheet": ws.title,
                        "financeRoleType": "nonTeaching" if teacher_name in NON_TEACHING_FINANCE_NAMES else "teaching",
                        "employmentStatus": "departed" if teacher_name in DEPARTED_FINANCE_ONLY_NAMES else "active",
                        "excludeFromTeacherDiff": teacher_name in DEPARTED_FINANCE_ONLY_NAMES,
                    }
                )

        if not revenue_section:
            continue
        revenue_name = text(ws.cell(row_idx, revenue_section["nameCol"]).value)
        if not revenue_name or revenue_name in {"合计", "总合计"}:
            continue
        revenue_fee = number(ws.cell(row_idx, revenue_section["feeCol"]).value, None)
        revenue_amount = number(ws.cell(row_idx, revenue_section["commissionCol"]).value, None)
        revenue_markers = []
        for col_idx in revenue_section["dateCols"]:
            parsed = parse_lesson_marker(ws.cell(row_idx, col_idx).value)
            if parsed:
                parsed["cell"] = ws.cell(row_idx, col_idx).coordinate
                revenue_markers.append(parsed)
        revenue_records.append(
            {
                "period": "2026-05",
                "teacherName": teacher_name,
                "teacherSheet": ws.title,
                "studentName": revenue_name,
                "courseFee": revenue_fee,
                "amount": revenue_amount if revenue_amount is not None else revenue_fee,
                "sourceSection": "revenue",
                "sourceRow": row_idx,
                "lessonMarkerCount": len(revenue_markers),
                "rawMarkers": [marker["raw"] for marker in revenue_markers],
            }
        )
        for marker in revenue_markers:
            revenue_attendance.append(
                {
                    "teacherName": teacher_name,
                    "teacher": teacher_name,
                    "studentName": revenue_name,
                    "courseFee": revenue_fee,
                    "amount": revenue_amount if revenue_amount is not None else revenue_fee,
                    "date": marker.get("date", ""),
                    "period": marker.get("date", "")[:7] if marker.get("date") else "",
                    "lessonNo": marker.get("lessonNo"),
                    "raw": marker.get("raw", ""),
                    "cell": marker.get("cell", ""),
                    "status": "待核收入" if not marker.get("needsReview") else "待复核",
                    "needsReview": marker.get("needsReview", False),
                    "reason": marker.get("reason", ""),
                    "sourceSection": "revenue",
                    "sourceWorkbook": MAY_SALARY_PATH.name,
                    "sourceSheet": ws.title,
                    "financeRoleType": "nonTeaching" if teacher_name in NON_TEACHING_FINANCE_NAMES else "teaching",
                    "employmentStatus": "departed" if teacher_name in DEPARTED_FINANCE_ONLY_NAMES else "active",
                }
            )

    return payout_records, payout_attendance, revenue_records, revenue_attendance


def load_may_workbook_payload():
    workbook = load_workbook(MAY_SALARY_PATH, data_only=True, read_only=True)
    summary_rows = extract_summary_rows(workbook["小课产值"])
    expense_rows = extract_expense_rows(workbook["日常消费表"])
    payout_records = []
    payout_attendance = []
    revenue_records = []
    revenue_attendance = []
    for ws in workbook.worksheets:
        if ws.title in IGNORE_SHEETS:
            continue
        records, attendance, revenue_rows, revenue_att_rows = extract_teacher_sheet(ws)
        payout_records.extend(records)
        payout_attendance.extend(attendance)
        revenue_records.extend(revenue_rows)
        revenue_attendance.extend(revenue_att_rows)
    return {
        "summaryRows": summary_rows,
        "expenseRows": expense_rows,
        "payoutRecords": payout_records,
        "payoutAttendance": payout_attendance,
        "revenueRecords": revenue_records,
        "revenueAttendance": revenue_attendance,
    }


def normalize_key(teacher: str, student: str, date: str) -> tuple[str, str, str]:
    return (
        text(teacher).replace(" ", "").replace("　", ""),
        text(student).replace(" ", "").replace("　", ""),
        text(date),
    )


def student_context(raw: str, student: str) -> str:
    compact_raw = text(raw).replace(" ", "").replace("　", "")
    compact_student = text(student).replace(" ", "").replace("　", "")
    if not compact_raw or not compact_student:
        return compact_raw
    pos = compact_raw.find(compact_student)
    return compact_raw[pos : pos + 55] if pos >= 0 else compact_raw


def normalize_category(lesson_type: str, raw_text: str, student_name: str = "") -> str:
    compact = f"{student_context(raw_text, student_name)} {lesson_type or ''} {raw_text or ''}"
    compact = compact.replace("（", "(").replace("）", ")").replace(" ", "").replace("　", "")
    if "试听" in compact:
        return "试听"
    if "免费" in compact or "送课" in compact:
        return "免费课"
    if "班课转出" in compact or "转出" in compact:
        return "班课转出"
    if "加课" in compact:
        return "加课一对一" if "一对一" in compact else "加课"
    if "一对一" in compact:
        return "一对一"
    if "小班课" in compact or "小班" in compact:
        return "小班课"
    if "托管" in compact:
        return "托管"
    if "刷题" in compact or "真题卷" in compact or "出门测" in compact:
        return "刷题/测试"
    if "科学" in compact:
        return "科学课"
    if re.search(r"补.*班.*春季.*第\d+次课", compact) or "补课" in compact:
        return "春季班补课"
    if "补" in compact:
        return "补课/待确认"
    return "未分类"


def schedule_row_label(row: dict, student_name: str = "") -> str:
    category = normalize_category(row.get("lessonType", ""), row.get("rawText", ""), student_name)
    return (
        f"{row['sourceFile']} {row.get('sourceSheet') or row.get('sheet') or ''}!{row['cell']} "
        f"{row['startTime']}-{row['endTime']} {category} [原备注:{row.get('lessonType') or '-'}]"
    )


def salary_row_label(row: dict) -> str:
    return (
        f"{row['teacherName']}!{row['cell']} {row['raw']} "
        f"课时费={row.get('courseFee')} 提成={row.get('commission')}"
    )


def priority_for(status: str, categories: list[str]) -> str:
    if status == "工资表有/排课无":
        return "高"
    if status == "匹配":
        return "已核对"
    if any(cat in {"小班课", "一对一", "加课", "加课一对一"} for cat in categories):
        return "高"
    if any(cat in {"班课转出", "春季班补课", "补课/待确认"} for cat in categories):
        return "中"
    if any(cat in {"试听", "免费课", "刷题/测试", "托管"} for cat in categories):
        return "低"
    return "中"


def build_may_schedule_sessions():
    schedule_payload = json.loads(SCHEDULE_JSON_PATH.read_text(encoding="utf-8"))
    sessions = []
    for row in schedule_payload["records"]:
        if not str(row.get("date", "")).startswith("2026-05"):
            continue
        if row.get("needsReview"):
            continue
        sessions.append(
            {
                "sourceId": f"{row['sourceFile']}|{row['sheet']}|{row['cell']}",
                "sourceFile": row["sourceFile"],
                "sourceSheet": row["sheet"],
                "cell": row["cell"],
                "teacherName": row["teacher"],
                "date": row["date"],
                "period": row["date"][:7],
                "startTime": row["startTime"],
                "endTime": row["endTime"],
                "hours": row["durationHours"],
                "studentNames": row.get("studentNames") or [],
                "studentCount": row.get("studentCount") or 0,
                "lessonTypeRaw": row.get("lessonType") or "",
                "rawText": row.get("rawText") or "",
                "status": "待对账",
                "importConfidence": "自动可信",
            }
        )
    return sessions


def build_may_reconciliation(schedule_sessions: list[dict], payout_rows: list[dict]):
    schedule_index = defaultdict(list)
    for row in schedule_sessions:
        for student_name in row.get("studentNames") or []:
            schedule_index[normalize_key(row["teacherName"], student_name, row["date"])].append(row)
    salary_index = defaultdict(list)
    for row in payout_rows:
        if row.get("needsReview") or row.get("period") != "2026-05":
            continue
        if row.get("employmentStatus") == "departed":
            continue
        salary_index[normalize_key(row["teacherName"], row["studentName"], row["date"])].append(row)

    rows = []
    for item in sorted(set(schedule_index) | set(salary_index)):
        teacher_name, student_name, date = item
        s_rows = schedule_index.get(item, [])
        w_rows = salary_index.get(item, [])
        if s_rows and w_rows:
            status = "匹配"
        elif s_rows:
            status = "排课有/工资表无"
        else:
            status = "工资表有/排课无"
        categories = [normalize_category(r.get("lessonTypeRaw", ""), r.get("rawText", ""), student_name) for r in s_rows]
        if not categories and w_rows:
            categories = ["工资表孤立记录"]
        issue = {
            "priority": priority_for(status, categories),
            "status": status,
            "category": " / ".join(sorted(set(categories))),
            "teacherName": teacher_name,
            "studentName": student_name,
            "date": date,
            "scheduleSource": "；".join(schedule_row_label(row, student_name) for row in s_rows),
            "salarySource": "；".join(salary_row_label(row) for row in w_rows),
        }
        rows.append(issue)
    return rows


def build_teacher_precheck(schedule_sessions: list[dict], payout_rows: list[dict], reconciliation_rows: list[dict]):
    teacher_summary = defaultdict(
        lambda: {
            "teacherName": "",
            "period": "2026-05",
            "scheduleStudentSessions": 0,
            "scheduleHours": 0.0,
            "salaryMarkers": 0,
            "matchedMarkers": 0,
            "scheduleOnly": 0,
            "salaryOnly": 0,
            "highPriorityDifferences": 0,
        }
    )
    for session in schedule_sessions:
        item = teacher_summary[session["teacherName"]]
        item["teacherName"] = session["teacherName"]
        item["scheduleStudentSessions"] += len(session.get("studentNames") or [])
        item["scheduleHours"] += to_number(session.get("hours")) * max(1, len(session.get("studentNames") or []))
    for row in payout_rows:
        if row.get("needsReview") or row.get("period") != "2026-05":
            continue
        if row.get("employmentStatus") == "departed":
            continue
        item = teacher_summary[row["teacherName"]]
        item["teacherName"] = row["teacherName"]
        item["salaryMarkers"] += 1
    for row in reconciliation_rows:
        item = teacher_summary[row["teacherName"]]
        item["teacherName"] = row["teacherName"]
        if row["status"] == "匹配":
            item["matchedMarkers"] += 1
        elif row["status"] == "排课有/工资表无":
            item["scheduleOnly"] += 1
        elif row["status"] == "工资表有/排课无":
            item["salaryOnly"] += 1
        if row["priority"] == "高" and row["status"] != "匹配":
            item["highPriorityDifferences"] += 1
    result = []
    for item in teacher_summary.values():
        item["scheduleHours"] = round(item["scheduleHours"], 2)
        result.append(item)
    return sorted(result, key=lambda row: row["teacherName"])


def build_may_teacher_questions(reconciliation_rows: list[dict]) -> str:
    lines = [
        "## 2026年5月补充核查",
        "",
        "这部分是五月真实工资表接入后的第一轮试算，先按高优先级问题核对。",
        "",
        "### 工资表有，但排课表没有",
    ]
    isolated = [row for row in reconciliation_rows if row["status"] == "工资表有/排课无"]
    if not isolated:
        lines.append("- 当前没有“工资表有但排课没有”的孤立课次。")
    else:
        for row in isolated[:20]:
            lines.append(
                f"- {row['date']} {row['teacherName']} {row['studentName']}：工资表写了 {row['salarySource']}，"
                "请确认排课表里是否漏填、姓名不一致，还是这节课不在该老师排课表中。"
            )
    lines.extend(["", "### 排课有，但工资表没有的高优先级记录"])
    high_rows = [row for row in reconciliation_rows if row["priority"] == "高" and row["status"] == "排课有/工资表无"]
    if not high_rows:
        lines.append("- 当前没有五月高优先级缺薪记录。")
    else:
        by_teacher = defaultdict(list)
        for row in high_rows:
            by_teacher[row["teacherName"]].append(row)
        for teacher_name in sorted(by_teacher):
            lines.append("")
            lines.append(f"#### {teacher_name}")
            for row in by_teacher[teacher_name][:10]:
                lines.append(f"- {row['date']} {row['studentName']} [{row['category']}]：{row['scheduleSource']}")
            if len(by_teacher[teacher_name]) > 10:
                lines.append(f"- 另有 {len(by_teacher[teacher_name]) - 10} 条同类记录，建议按同口径一次核完。")
    return "\n".join(lines)


def build_finance_readiness_text(may_payload: dict, may_reconciliation: list[dict]) -> str:
    summary_rows = may_payload["summaryRows"]
    expense_rows = may_payload["expenseRows"]
    payout_rows = may_payload["payoutAttendance"]
    revenue_rows = may_payload["revenueAttendance"]
    base_salary_total = round(sum(to_number(row.get("基础工资")) for row in summary_rows), 2)
    social_total = round(sum(to_number(row.get("社保")) for row in summary_rows), 2)
    commission_total = round(sum(to_number(row.get("课时提成")) for row in summary_rows), 2)
    makeup_total = round(sum(to_number(row.get("补课提成")) for row in summary_rows), 2)
    lesson_income_total = round(sum(to_number(row.get("amount")) for row in revenue_rows if row.get("period") == "2026-05"), 2)
    expense_total = round(sum(to_number(row.get("amount")) for row in expense_rows), 2)
    return "\n".join(
        [
            "# 5月真实工资表预导入审计",
            "",
            "## 当前结论",
            f"- 已读取 `{MAY_SALARY_PATH.name}` 的 `小课产值`、`日常消费表` 和各老师工资明细。",
            "- 当前仍是试算口径，不直接改正式工资和正式课销。",
            f"- 五月排课已接入 {sum(1 for row in may_payload['scheduleSessions'])} 节可信课次，可与五月工资表逐条比对。",
            "",
            "## 试算基础",
            f"- 老师汇总行：{len(summary_rows)} 人",
            f"- 课时工资标记：{len([row for row in payout_rows if row.get('period') == '2026-05'])} 条",
            f"- 课时收入标记：{len([row for row in revenue_rows if row.get('period') == '2026-05'])} 条",
            f"- 日常成本明细：{len(expense_rows)} 条",
            f"- 五月课时收入试算：{lesson_income_total:.2f}",
            f"- 五月课时提成试算：{commission_total:.2f}",
            f"- 五月补课提成试算：{makeup_total:.2f}",
            f"- 五月基础工资合计：{base_salary_total:.2f}",
            f"- 五月社保合计：{social_total:.2f}",
            f"- 五月日常成本合计：{expense_total:.2f}",
            "",
            "## 需要继续核的地方",
            f"- 五月排课/工资差异：{sum(1 for row in may_reconciliation if row['status'] != '匹配')} 条，"
            f"其中高优先级 {sum(1 for row in may_reconciliation if row['priority'] == '高' and row['status'] != '匹配')} 条。",
            "- 周珊按学管 / 排课岗处理，保留工资成本，不进入教学老师差异口径。",
            "- 万盈盈按已离职历史结算处理，保留离职前工资结算，不再作为在岗教学老师参与差异展示。",
            "- 五月费用按分类栏抽取，没有把左上角的汇总流水块重复计入，避免重复放大成本。",
        ]
    )


def build_bundle():
    base_bundle = json.loads(BASE_BUNDLE_PATH.read_text(encoding="utf-8"))
    may_payload = load_may_workbook_payload()
    may_schedule_sessions = build_may_schedule_sessions()
    may_reconciliation = build_may_reconciliation(may_schedule_sessions, may_payload["payoutAttendance"])
    may_teacher_precheck = build_teacher_precheck(may_schedule_sessions, may_payload["payoutAttendance"], may_reconciliation)

    bundle = deepcopy(base_bundle)
    bundle["generatedAt"] = "2026-06-23"
    bundle["scheduleSessions"] = sorted(
        [*bundle.get("scheduleSessions", []), *may_schedule_sessions],
        key=lambda row: (row.get("date", ""), row.get("teacherName", ""), row.get("sourceId", "")),
    )
    bundle["salaryAttendance"] = sorted(
        [*bundle.get("salaryAttendance", []), *may_payload["payoutAttendance"]],
        key=lambda row: (row.get("date", ""), row.get("teacherName", ""), row.get("studentName", ""), row.get("cell", "")),
    )
    bundle["salaryRevenueAttendance"] = sorted(
        [*(bundle.get("salaryRevenueAttendance") or []), *may_payload["revenueAttendance"]],
        key=lambda row: (row.get("date", ""), row.get("teacherName", ""), row.get("studentName", ""), row.get("cell", "")),
    )
    bundle["expenseRows"] = sorted(
        [*bundle.get("expenseRows", []), *may_payload["expenseRows"]],
        key=lambda row: (row.get("date", ""), row.get("category", ""), row.get("name", "")),
    )
    bundle["teacherFinanceSummaryRows"] = sorted(
        [*(bundle.get("teacherFinanceSummaryRows") or []), *may_payload["summaryRows"]],
        key=lambda row: (row.get("period", ""), row.get("teacherName", "")),
    )
    bundle["teacherMonthPrecheck"] = sorted(
        [*bundle.get("teacherMonthPrecheck", []), *may_teacher_precheck],
        key=lambda row: (row.get("period", ""), row.get("teacherName", "")),
    )
    bundle["reconciliationIssues"] = sorted(
        [*bundle.get("reconciliationIssues", []), *[row for row in may_reconciliation if row["status"] != "匹配"]],
        key=lambda row: (row.get("date", ""), row.get("teacherName", ""), row.get("studentName", ""), row.get("status", "")),
    )

    may_questions = build_may_teacher_questions(may_reconciliation)
    base_questions = text(bundle.get("teacherQuestionsText"))
    bundle["teacherQuestionsText"] = f"{base_questions}\n\n{may_questions}".strip()

    may_payload["scheduleSessions"] = may_schedule_sessions
    may_readiness = build_finance_readiness_text(may_payload, may_reconciliation)
    base_readiness = text(bundle.get("financeReadinessText"))
    bundle["financeReadinessText"] = f"{base_readiness}\n\n{may_readiness}".strip()

    by_teacher = sorted(bundle["teacherMonthPrecheck"], key=lambda row: (row.get("period", ""), row.get("teacherName", "")))
    by_schedule_period = Counter(row.get("period", "") for row in bundle["scheduleSessions"])
    issue_rows = bundle["reconciliationIssues"]
    bundle["summary"] = {
        "scheduleSessionCount": len(bundle["scheduleSessions"]),
        "scheduleStudentSessionCount": sum(len(row.get("studentNames") or []) for row in bundle["scheduleSessions"]),
        "salaryAttendanceCount": len(bundle["salaryAttendance"]),
        "expenseRowCount": len(bundle["expenseRows"]),
        "issueCount": len(issue_rows),
        "highPriorityIssueCount": sum(1 for row in issue_rows if row.get("priority") == "高"),
        "bySchedulePeriod": dict(sorted((period, count) for period, count in by_schedule_period.items() if period)),
        "byTeacher": by_teacher,
    }
    return bundle


def write_preimport_files(bundle: dict) -> None:
    serialized = json.dumps(bundle, ensure_ascii=False, separators=(",", ":"))
    content = (
        f"window.JRC_PREIMPORT_SUMMARY = {serialized};\n"
        "window.JRC_PREIMPORT_BUNDLE = window.JRC_PREIMPORT_BUNDLE || window.JRC_PREIMPORT_SUMMARY;\n"
    )
    PORTAL_PREIMPORT_PATH.write_text(content, encoding="utf-8")
    PUBLIC_PREIMPORT_PATH.write_text(content, encoding="utf-8")
    OUT_BUNDLE_PATH.write_text(json.dumps(bundle, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> int:
    bundle = build_bundle()
    write_preimport_files(bundle)
    print(PORTAL_PREIMPORT_PATH)
    print(PUBLIC_PREIMPORT_PATH)
    print(OUT_BUNDLE_PATH)
    print(json.dumps(bundle["summary"], ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
